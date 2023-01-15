import os

from openpyxl import load_workbook


class ExcelUtils:

    @staticmethod
    def getCellData(cell, sheet="Sheet1", path=None):

        if path is None:
            path = os.path.abspath("./").split("odysseus_tests")[0] + "odysseus_tests\\tests\\resources\\excels\\defaultExcelData.xlsx"

        wb = load_workbook(path)
        if sheet is not None and sheet != "" and sheet in wb.sheetnames:
            ws = wb[sheet]
            if cell is not None and cell != "":
                return ws[cell].value
            else:
                return "Cell not found"
        else:
            return "Sheet not found"

    @staticmethod
    def getRowColData(col, row, sheet="Sheet1", path=None):

        if path is None:
            path = os.path.abspath("./").split("odysseus_tests")[0] + "odysseus_tests\\tests\\resources\\excels\\defaultExcelData.xlsx"

        wb = load_workbook(path)
        if sheet is not None and sheet in wb.sheetnames:
            ws = wb[sheet]
            if row and col is not None:
                return ws[(str(col) + str(row))].value
            else:
                return "Cell not found"
        else:
            return "Sheet not found"

    @staticmethod
    def getColData(col, sheet="Sheet1", path=None):

        if path is None:
            path = os.path.abspath("./").split("odysseus_tests")[0] + "odysseus_tests\\tests\\resources\\excels\\defaultExcelData.xlsx"

        wb = load_workbook(path)
        if sheet is not None and sheet in wb.sheetnames:
            ws = wb[sheet]
            if col is not None:
                char = ""
                for i in ws[1]:
                    if i.value == col:
                        char = i.column_letter
                colVals = []
                for j in range(1, len(ws[char])):
                    colVals.append(ws[char][j].value)
                return colVals
            else:
                return "Cell not found"
        else:
            return "Sheet not found"

    @staticmethod
    def setData(path, sheet, cell):
        pass





